# %% [markdown]
# # Menu Controle de Despesas

# %%
from openpyxl.utils.dataframe import dataframe_to_rows
import xml.etree.ElementTree as ET
import datetime as dt
import openpyxl as xl
import pandas as pd
import regex as re
import time
import os


# %% [markdown]
# ### titulos

# %%
titulo_empresa = """
 ██████████                                                               
▒▒███▒▒▒▒▒█                                                               
 ▒███  █ ▒  █████████████   ████████  ████████   ██████   █████   ██████  
 ▒██████   ▒▒███▒▒███▒▒███ ▒▒███▒▒███▒▒███▒▒███ ███▒▒███ ███▒▒   ▒▒▒▒▒███ 
 ▒███▒▒█    ▒███ ▒███ ▒███  ▒███ ▒███ ▒███ ▒▒▒ ▒███████ ▒▒█████   ███████ 
 ▒███ ▒   █ ▒███ ▒███ ▒███  ▒███ ▒███ ▒███     ▒███▒▒▒   ▒▒▒▒███ ███▒▒███ 
 ██████████ █████▒███ █████ ▒███████  █████    ▒▒██████  ██████ ▒▒████████
▒▒▒▒▒▒▒▒▒▒ ▒▒▒▒▒ ▒▒▒ ▒▒▒▒▒  ▒███▒▒▒  ▒▒▒▒▒      ▒▒▒▒▒▒  ▒▒▒▒▒▒   ▒▒▒▒▒▒▒▒ 
                            ▒███                                          
                            █████                                         
                           ▒▒▒▒▒      
"""

registrar_titulo = '''
▀██▀▀█▄                    ██           ▄                   
 ██   ██    ▄▄▄▄    ▄▄▄ ▄ ▄▄▄   ▄▄▄▄  ▄██▄   ▄▄▄▄   ▄▄▄ ▄▄  
 ██▀▀█▀   ▄█▄▄▄██  ██ ██   ██  ██▄ ▀   ██   ▀▀ ▄██   ██▀ ▀▀ 
 ██   █▄  ██        █▀▀    ██  ▄ ▀█▄▄  ██   ▄█▀ ██   ██     
▄██▄  ▀█▀  ▀█▄▄▄▀  ▀████▄ ▄██▄ █▀▄▄█▀  ▀█▄▀ ▀█▄▄▀█▀ ▄██▄    
                  ▄█▄▄▄▄▀                                   
                                                            
'''

visualizar_titulo = '''
▀██▀  ▀█▀  ██                          ▀██   ██                          
 ▀█▄  ▄▀  ▄▄▄   ▄▄▄▄   ▄▄▄▄   ▄▄▄ ▄▄▄   ██  ▄▄▄  ▄▄▄▄▄▄   ▄▄▄▄   ▄▄▄ ▄▄  
  ██  █    ██  ██▄ ▀  ▀▀ ▄██   ██  ██   ██   ██  ▀  ▄█▀  ▀▀ ▄██   ██▀ ▀▀ 
   ███     ██  ▄ ▀█▄▄ ▄█▀ ██   ██  ██   ██   ██   ▄█▀    ▄█▀ ██   ██     
    █     ▄██▄ █▀▄▄█▀ ▀█▄▄▀█▀  ▀█▄▄▀█▄ ▄██▄ ▄██▄ ██▄▄▄▄█ ▀█▄▄▀█▀ ▄██▄                                                                         

'''

alterar_titulo = '''
    █     ▀██    ▄                                   
   ███     ██  ▄██▄    ▄▄▄▄  ▄▄▄ ▄▄   ▄▄▄▄   ▄▄▄ ▄▄  
  █  ██    ██   ██   ▄█▄▄▄██  ██▀ ▀▀ ▀▀ ▄██   ██▀ ▀▀ 
 ▄▀▀▀▀█▄   ██   ██   ██       ██     ▄█▀ ██   ██     
▄█▄  ▄██▄ ▄██▄  ▀█▄▀  ▀█▄▄▄▀ ▄██▄    ▀█▄▄▀█▀ ▄██▄    
                                                     
                                                     
'''

# %% [markdown]
# ## Variaveis globais
# 

# %%
debug = True
TabList = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
today = dt.date.today()
current_month_index = today.month - 1
dados = pd.DataFrame()
tag_padrao = '{http://www.portalfiscal.inf.br/nfe}'

csv_file = './src/Data/Dados.csv'
excel_file = './src/template/DM-2025 - Template.xlsx'
if debug:
    save_csv_file = 'src/Data/Dados_Debug.csv'
else:
    save_csv_file = './src/Data/Dados.csv'
output_file = os.path.join(os.path.expanduser('~'), 'Downloads', f'Gastos-{today.year}-{today.month:02d}.xlsx')

# %% [markdown]
# ### Carregar Dados

# %%
def carregar_dados():
    '''
    Carrega os dados do arquivo CSV para o DataFrame pandas
    '''
    global dados
    dados = pd.read_csv(csv_file)

# %% [markdown]
# ## Menu Inicial

# %%
def exibir_titulo(texto):
    '''Exibe o titulo estilizado na tela'''
    os.system('cls')
    print(texto)

def finalizar_app():
    '''Exibe mensagem de finalizacao do app'''
    exibir_subtitulo('Finalizando app')

def voltar_ao_menu_principal():
    '''
    Solicita uma tecla e retorna ao menu principal
        
    Outputs:
    - Retorna ao menu principal
    '''
    input('\nDigite uma tecla para voltar ao menu ')
    main_menu()

def opcao_invalida():
    '''Case o usuario insira um input invalido, retorne ele ao menu principal
    
    Outputs:
    - Retorna ao menu principal
    '''
    print('Input inválido! Retornando ao menu...')
    voltar_ao_menu_principal()

def exibir_subtitulo(texto):
    '''Limpa a tela e exibe o subtitulo estilizado na tela
    
    Inputs:
    - texto: str - o texto do subtitulo
    '''
    os.system('cls')
    linha = '*' * (len(texto))
    print(linha)
    print(texto)
    print(linha)
    print()

# %% [markdown]
# ## Paginacao

# %%

def exibir_pagina(tabela, numero, max):
    '''
    Exibe ao usuario parte da lista de jogadoras, baseado no numero inserido (numero da pagina).
    Caso o ultimo index a ser exibido passe do valor maximo de jogadora registradas, o ultimo index é alterado para o limite.
    '''
    num_min = (numero - 1) * 10
    num_max = num_min + 10
    if num_max > max:
        num_max = max
    print(tabela[num_min:num_max])

def converter_input_numero_pagina(input_numero_pagina):
    '''
    Converte o input do usuario para um numero inteiro.
    Caso o input seja vazio ou invalido, retorna None
    '''
    try:
        input_numero_pagina = int(input_numero_pagina)
        return input_numero_pagina
    except:
        return None

def paginacao(tabela, subtitulo=str):
    '''
    Função que implementa a paginação para exibir a tabela em partes para o usuario.
    '''
    exibir_subtitulo(subtitulo)
    
    input_numero_pagina = 1
    pagina_max = len(tabela)
    numero_de_paginas = int(((pagina_max - 1)/10)+1)

    exibir_pagina(tabela, input_numero_pagina, pagina_max)
    print(f"\n-----------! Numeor de paginas: {numero_de_paginas} !----------")
    
    input_numero_pagina = converter_input_numero_pagina(input("\nInsira o numero da pagina que deseja visualizar ou pressione enter para sair: "))
    
    while isinstance(input_numero_pagina, int):
        if isinstance(input_numero_pagina, int):
            if 1 <= input_numero_pagina <= numero_de_paginas:
                exibir_subtitulo(subtitulo)
                exibir_pagina(tabela, input_numero_pagina, pagina_max)
                print(f"\n-----------! Numero de paginas: {numero_de_paginas} !----------")
                input_numero_pagina = converter_input_numero_pagina(input("\nInsira o numero da pagina que deseja visualizar ou pressione enter para sair: "))
            else:
                exibir_subtitulo(subtitulo)
                print("Numero de pagina inserido invalido.")
                print(f"\n-----------! Numero de paginas: {numero_de_paginas} !----------")
                input_numero_pagina = converter_input_numero_pagina(input("\nInsira o numero da pagina que deseja visualizar ou pressione enter para sair: "))
        else:
            print("\nRetornando ao menu principal...\n")
            voltar_ao_menu_principal()
    
def visualizar_dados():
    '''Lista jogadoras registradas no banco de dados
    
    Inputs:
    - Dictionary de jogadoras
    
    Outputs:
    - Exibe lista de jogadoras na tela
    
    '''
    global dados

    paginacao(dados, "Lista de Gastos")

    voltar_ao_menu_principal()

# %% [markdown]
# # Menus

# %% [markdown]
# ## Registrar

# %%
regex_pattern_date01 = r'[0-9]{4}-[0-9]{2}-[0-9]{2}'
regex_pattern_date04 = r'[0-9]{4}/[0-9]{2}/[0-9]{2}'
regex_pattern_date02 = r'[0-9]{2}/[0-9]{2}/[0-9]{4}'
regex_pattern_date03 = r'[0-9]{2}-[0-9]{2}-[0-9]{4}'

dias_da_semana = {
    0: 'Segunda-feira',
    1: 'Terça-feira',
    2: 'Quarta-feira',
    3: 'Quinta-feira',
    4: 'Sexta-feira',
    5: 'Sábado',
    6: 'Domingo'
}

def request_date(texto):
    date_input = input(f"Insira a data de {texto} (no formato AAAA-MM-DD) ou enter para sair: ")

    try:
        if re.fullmatch(regex_pattern_date01, date_input):
            print(f"Data '{date_input}' está no formato AAAA-MM-DD.")
            data = dt.datetime.strptime(date_input, "%Y-%m-%d").date()
        elif re.fullmatch(regex_pattern_date04, date_input):
            print(f"Data '{date_input}' está no formato AAAA/MM/DD.")
            data = dt.datetime.strptime(date_input, "%Y/%m/%d").date()
        elif re.fullmatch(regex_pattern_date02, date_input):
            print(f"Data '{date_input}' está no formato DD/MM/AAAA.")
            data = dt.datetime.strptime(date_input, "%d/%m/%Y").date()
        elif re.fullmatch(regex_pattern_date03, date_input):
            print(f"Data '{date_input}' está no formato DD-MM-AAAA.")
            data = dt.datetime.strptime(date_input, "%d-%m-%Y").date()
        elif date_input == '':
            return None
        else:
            print("Valor inserido Invalido. Tente novamente.\n")
            return request_date(texto)
    except ValueError as e:
        print('Valor inserido invalido. Tente novamente.\n')
        print(f"Error parsing date: {e}")
        if date_input != '':
            return request_date(texto)
        else:
            return None
    else:
        if debug:
            weekday = dias_da_semana[data.weekday()]
            print(f"DEBUG: O dia da semana é: {weekday}")
        return data

# %%
def registrar_gasto(tipo, data_emissao, fornecedor, data_pagamento, nNF, valor, valor_icms, valor_cofins, valor_pis, valor_ipi):
    novo_gasto = {
        'Tipo': tipo,
        'Data de emissão': data_emissao,
        'Fornecedor': fornecedor,
        'Data de Vencimento': data_pagamento,
        'N° da NF': nNF,
        'V. Total da NF': valor,
        'ICMS': valor_icms,
        'COFINS': valor_cofins,
        'PIS': valor_pis,
        'IPI': valor_ipi
    }
    global dados
    dados = pd.concat([dados, pd.DataFrame([novo_gasto])], ignore_index=True)
    dados.to_csv(save_csv_file, index=False)
    print("Gasto registrado com sucesso!")

def input_valor_monetario(prompt):
    valor_str = 'valor'
    while True and valor_str != '':
        valor_str = input(prompt).replace(',', '.')
        if re.match(r'^\d+(\.\d{1,2})?$', valor_str):
            return float(valor_str)
        else:
            print("Valor inválido. Por favor, insira um valor monetário válido (ex: 1234.56). Ou aperte Enter para sair.")
    
    return ''

def registro_manual(tipo):
    tipo = tipo.upper()
    print(f"\nRegistrando gasto do tipo: {tipo}")
    fornecedor = input("Fornecedor: ")
    if fornecedor == '':
        print("Registro cancelado.")
        return
    data_emissao = request_date("Emissão")
    if data_emissao is None:
        print("Registro cancelado.")
        return
    data_pagamento = request_date("Vencimento")
    if data_pagamento is None:
        print("Registro cancelado.")
        return
    try:
        nNF = int(input("N° da NF: "))
    except ValueError:
        print("Número da NF inválido. Registro cancelado.")
        return
    valor = input_valor_monetario("V. Total da NF: R$ ")
    if valor == '':
        print("Registro cancelado.")
        return
    valor_icms = input_valor_monetario("ICMS: R$ ")
    if valor_icms == '':
        print("Registro cancelado.")
        return
    valor_cofins = input_valor_monetario("COFINS: R$ ")
    if valor_cofins == '':
        print("Registro cancelado.")
        return
    valor_pis = input_valor_monetario("PIS: R$ ")
    if valor_pis == '':
        print("Registro cancelado.")
        return
    valor_ipi = input_valor_monetario("IPI: R$ ")
    if valor_ipi == '':
        print("Registro cancelado.")
        return

    registrar_gasto(tipo, data_emissao, fornecedor, data_pagamento, nNF, valor, valor_icms, valor_cofins, valor_pis, valor_ipi)

# %%
def registro_automatico():
    for file in os.listdir('./src/Data/xml'):
        if file.endswith('.xml'):
            caminho_arquivo = os.path.join('./src/Data/xml', file)
            tree = ET.parse(caminho_arquivo)
            root = tree.getroot()
            base = root.find(f'{tag_padrao}NFe').find(f'{tag_padrao}infNFe')

            tipo = 'C'
            dataE = base.find(f'{tag_padrao}ide').find(f'{tag_padrao}dhEmi').text
            dataE = dataE.split('T')[0]
            fornecedor = base.find(f'{tag_padrao}emit').find(f'{tag_padrao}xNome').text
            if base.find(f'{tag_padrao}cobr') is None:
                dataV = dataE
            else:
                dataV = base.find(f'{tag_padrao}cobr').find(f'{tag_padrao}dup').find(f'{tag_padrao}dVenc').text
            nNF = int(base.find(f'{tag_padrao}ide').find(f'{tag_padrao}nNF').text)
            valor_total = float(base.find(f'{tag_padrao}total').find(f'{tag_padrao}ICMSTot').find(f'{tag_padrao}vNF').text)
            valor_icms = float(base.find(f'{tag_padrao}total').find(f'{tag_padrao}ICMSTot').find(f'{tag_padrao}vICMS').text)
            valor_cofins = float(base.find(f'{tag_padrao}total').find(f'{tag_padrao}ICMSTot').find(f'{tag_padrao}vCOFINS').text)
            valor_pis = float(base.find(f'{tag_padrao}total').find(f'{tag_padrao}ICMSTot').find(f'{tag_padrao}vPIS').text)
            valor_ipi = float(base.find(f'{tag_padrao}total').find(f'{tag_padrao}ICMSTot').find(f'{tag_padrao}vIPI').text)

            print(f'Arquivo: {file} -> Nota Fiscal: {nNF} - Fornecedor: {fornecedor} - Valor Total: {valor_total}')

            registrar_gasto(tipo, dataE, fornecedor, dataV, nNF, valor_total, valor_icms, valor_cofins, valor_pis, valor_ipi)

    print('\nRegistro automatico finalizado.')



# %%
def exibir_opcoes_registrar():
    '''Exibe todas as opcoes de input do usuario do menu registrar dados'''
    print('1. Compras - Registro Automatico - via Arquivos XML')
    print('2. Despesas - Registro Manual')
    print('3. Servicos - Registro Manual')
    print('9. Voltar ao menu Principal\n')

def escolher_opcao_registrar():
    ''' Solicita a executa a opcao escolhida pelo usuario
    
    Outputs:
    -Executa a opcao escolhida pelo usuario
    '''
    try:
        opcao_escolhida = int(input('Escolha uma opção: '))
        # opcao_escolhida = int(opcao_escolhida)

        if opcao_escolhida == 1: 
            registro_automatico()
            input('\nDigite uma tecla para voltar ao menu ')
            menu_registrar_dados()
        elif opcao_escolhida == 2: 
            registro_manual('D')
            input('\nDigite uma tecla para voltar ao menu ')
            menu_registrar_dados()
        elif opcao_escolhida == 3: 
            registro_manual('S')
            input('\nDigite uma tecla para voltar ao menu ')
            menu_registrar_dados()
        elif opcao_escolhida == 9: 
            voltar_ao_menu_principal()
        else: 
            opcao_invalida()
    except Exception as e:
        if debug:
            print(f'Error: {type(e).__name__}')
            print(f"Error message: {e}")
        opcao_invalida()

def menu_registrar_dados():
    '''Menu para registrar dados de jogadoras'''
    exibir_titulo(registrar_titulo)
    exibir_opcoes_registrar()
    escolher_opcao_registrar()

# %% [markdown]
# ## Visualizar

# %%
def merge_csv_to_excel(csv_file, excel_file, output_file):
    # Load the existing Excel workbook
    workbook = xl.load_workbook(excel_file)
    sheet = workbook[TabList[current_month_index]]

    style = xl.styles.NamedStyle(name="standard_style")
    style.font = xl.styles.Font(name='Aptos', size=11)
    style.alignment = xl.styles.Alignment(horizontal='center', vertical='center')
    style.border = xl.styles.Border(
        left=xl.styles.Side(border_style='medium', color='000000'),
        right=xl.styles.Side(border_style='medium', color='000000'),
        top=xl.styles.Side(border_style='none', color='000000'),
        bottom=xl.styles.Side(border_style='none', color='000000')
    )
    workbook.add_named_style(style)

    # Read the CSV file into a DataFrame
    df = pd.read_csv(csv_file)

    # Append the DataFrame to the Excel sheet
    for r in dataframe_to_rows(df, index=False, header=False):
        if 'current_row' not in locals():
            current_row = 3
        for col_idx, value in enumerate(r, start=1):
            sheet.cell(row=current_row, column=col_idx, value=value)
            sheet.cell(row=current_row, column=col_idx).style = "standard_style"
            if col_idx in [ 5, 7, 8, 9, 10 ]:  # Columns with monetary values
                sheet.cell(row=current_row, column=col_idx).number_format = 'R$ #,##0.00'
            
        current_row += 1

    workbook.active = workbook[TabList[current_month_index]]
    # Save the updated workbook to a new file
    workbook.save(output_file)

# %%
def salvar_arquivo_excel():
    '''Salva o arquivo excel com os dados coletados'''
    try:
        merge_csv_to_excel(save_csv_file, excel_file, output_file)
    except Exception as e:
        print(f'Erro ao salvar: {e}')
        voltar_ao_menu_principal()
        return

    timeout = 2.0      # segundos máximos para esperar
    interval = 0.2     # intervalo entre checagens
    waited = 0.0
    while not os.path.isfile(output_file) and waited < timeout:
        time.sleep(interval)
        waited += interval

    if os.path.isfile(output_file):
        print(f'Arquivo salvo com sucesso em na pasta de Downloads como: {os.path.basename(output_file)}')
    else:
        print('Erro ao salvar o arquivo. Feche o Excel e tente novamente.')
    voltar_ao_menu_principal()

# %%
def filtrar_dados(tipo=None):
    global dados
    df_filtrado = dados.copy()

    if tipo:
        df_filtrado = df_filtrado[df_filtrado['Tipo'].str.contains(tipo, case=False, na=False)]

    return df_filtrado

# %%
def visualizar_por_categoria():
    '''Visualiza os dados filtrados por categoria'''
    tipo = input('Insira o tipo de despesa (C para Compras, D para Despesas, S para Serviços): ').strip().upper()
    dados_filtrados = filtrar_dados(tipo=tipo)
    paginacao(dados_filtrados, f"Dados Filtrados por Tipo: {tipo}")
    voltar_ao_menu_principal()

# %%
def exibir_opcoes_visualizar():
    '''Exibe todas as opcoes de input do usuario do menu registrar dados'''
    print('1. Visualizar todos os dados')
    print('2. Visualizar por Categorias')
    print('3. Salvar Arquivo Excel')
    print('9. Voltar ao menu Principal\n')
    
def escolher_opcao_visualizar():
    ''' Solicita a executa a opcao escolhida pelo usuario
    
    Outputs:
    -Executa a opcao escolhida pelo usuario
    '''
    try:
        opcao_escolhida = int(input('Escolha uma opção: '))
        # opcao_escolhida = int(opcao_escolhida)

        if opcao_escolhida == 1: 
            visualizar_dados()
        elif opcao_escolhida == 2: 
            visualizar_por_categoria()
        elif opcao_escolhida == 3: 
            salvar_arquivo_excel()
        elif opcao_escolhida == 9: 
            voltar_ao_menu_principal()
        else: 
            opcao_invalida()
    except Exception as e:
        if debug:
            print(f'Error: {type(e).__name__}')
            print(f"Error message: {e}")
        opcao_invalida()

def menu_visualizar_dados():
    '''Menu para registrar dados de jogadoras'''
    exibir_titulo(visualizar_titulo)
    exibir_opcoes_visualizar()
    escolher_opcao_visualizar()

# %% [markdown]
# ## Alterar

# %%
def alterar_por_nNF():
    '''Altera os dados de uma despesa baseado no N° da NF'''

    exibir_subtitulo("Alterar registro por N° da NF")
    nNF_input = input("Insira o N° da NF que deseja alterar: ")
    try:
        nNF_input = int(nNF_input)
    except ValueError:
        print("Número da NF inválido. Retornando ao menu...")
        menu_alterar_dados()
        return

    global dados
    if nNF_input not in dados['N° da NF'].values:
        print(f"Número da NF {nNF_input} não encontrado. Retornando ao menu...")
        menu_alterar_dados()
        return
    elif len(dados[dados['N° da NF'] == nNF_input]) > 1:
        print(f"Mais de um registro encontrado para a NF {nNF_input}.\n")
        print(dados[dados['N° da NF'] == nNF_input])
        indice = input("Insira o índice do registro que deseja apagar (conforme exibido acima): ")
        try:
            indice = int(indice)
            if indice not in dados.index:
                print("Índice inválido. Retornando ao menu...")
                menu_alterar_dados()
                return
        except ValueError:
            print("Valor inserido inválido. Retornando ao menu...")
            menu_alterar_dados()
            return
    else:
        indice = dados.index[dados['N° da NF'] == nNF_input][0]

    print(f"Alterando dados para a NF {nNF_input}:")
    
    novo_fornecedor = input(f"Fornecedor atual ({dados.at[indice, 'Fornecedor']}), novo valor (pressione Enter para manter): ")
    if novo_fornecedor:
        dados.at[indice, 'Fornecedor'] = novo_fornecedor

    nova_data_emissao = request_date("Emissão (pressione Enter para manter)")
    if nova_data_emissao:
        dados.at[indice, 'Data de emissão'] = nova_data_emissao

    nova_data_pagamento = request_date("Vencimento (pressione Enter para manter)")
    if nova_data_pagamento:
        dados.at[indice, 'Data de Vencimento'] = nova_data_pagamento

    novo_valor = input_valor_monetario(f"V. Total da NF atual (R$ {dados.at[indice, 'V. Total da NF']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor != '':
        dados.at[indice, 'V. Total da NF'] = novo_valor

    novo_valor_icms = input_valor_monetario(f"ICMS atual (R$ {dados.at[indice, 'ICMS']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_icms != '':
        dados.at[indice, 'ICMS'] = novo_valor_icms

    novo_valor_cofins = input_valor_monetario(f"COFINS atual (R$ {dados.at[indice, 'COFINS']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_cofins != '':
        dados.at[indice, 'COFINS'] = novo_valor_cofins

    novo_valor_pis = input_valor_monetario(f"PIS atual (R$ {dados.at[indice, 'PIS']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_pis != '':
        dados.at[indice, 'PIS'] = novo_valor_pis

    novo_valor_ipi = input_valor_monetario(f"IPI atual (R$ {dados.at[indice, 'IPI']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_ipi != '':
        dados.at[indice, 'IPI'] = novo_valor_ipi

    if novo_fornecedor or nova_data_emissao or nova_data_pagamento or novo_valor != '' or novo_valor_icms != '' or novo_valor_cofins != '' or novo_valor_pis != '' or novo_valor_ipi != '':
        dados.to_csv(save_csv_file, index=False)
        print("Dados alterados com sucesso!")
    else:
        print("Nenhum dado foi alterado.")
    
    input('\nDigite uma tecla para voltar ao menu.')
    menu_alterar_dados()



# %%
def alterar_por_data_emissao():
    '''Altera os dados de uma despesa baseado na Data de Emissão'''

    exibir_subtitulo("Alterar registro por Data de Emissão")
    
    data_emissao = request_date("Emissão para busca")
    if data_emissao is None:
        print("Data inválida. Retornando ao menu...")
        menu_alterar_dados()
        return

    global dados
    dados_filtrados = dados[dados['Data de emissão'] == str(data_emissao)]

    if dados_filtrados.empty:
        print(f"Nenhum registro encontrado para a data de emissão {data_emissao}. Retornando ao menu...")
        menu_alterar_dados()
        return

    print(f"Registros encontrados para a data de emissão {data_emissao}:")
    print(dados_filtrados)

    nNF_input = input("Insira o N° da NF que deseja alterar: ")
    try:
        nNF_input = int(nNF_input)
    except ValueError:
        print("Número da NF inválido. Retornando ao menu...")
        menu_alterar_dados()
        return

    if nNF_input not in dados_filtrados['N° da NF'].values:
        print(f"Número da NF {nNF_input} não encontrado na data de emissão {data_emissao}. Retornando ao menu...")
        menu_alterar_dados()
        return

    indice = dados.index[dados['N° da NF'] == nNF_input][0]
    print(f"Alterando dados para a NF {nNF_input}:")
    
    novo_fornecedor = input(f"Fornecedor atual ({dados.at[indice, 'Fornecedor']}), novo valor (pressione Enter para manter): ")
    if novo_fornecedor:
        dados.at[indice, 'Fornecedor'] = novo_fornecedor

    nova_data_emissao = request_date("Emissão (pressione Enter para manter)")
    if nova_data_emissao:
        dados.at[indice, 'Data de emissão'] = nova_data_emissao

    nova_data_pagamento = request_date("Vencimento (pressione Enter para manter)")
    if nova_data_pagamento:
        dados.at[indice, 'Data de Vencimento'] = nova_data_pagamento

    novo_valor = input_valor_monetario(f"V. Total da NF atual (R$ {dados.at[indice, 'V. Total da NF']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor != '':
        dados.at[indice, 'V. Total da NF'] = novo_valor

    novo_valor_icms = input_valor_monetario(f"ICMS atual (R$ {dados.at[indice, 'ICMS']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_icms != '':
        dados.at[indice, 'ICMS'] = novo_valor_icms

    novo_valor_cofins = input_valor_monetario(f"COFINS atual (R$ {dados.at[indice, 'COFINS']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_cofins != '':
        dados.at[indice, 'COFINS'] = novo_valor_cofins

    novo_valor_pis = input_valor_monetario(f"PIS atual (R$ {dados.at[indice, 'PIS']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_pis != '':
        dados.at[indice, 'PIS'] = novo_valor_pis

    novo_valor_ipi = input_valor_monetario(f"IPI atual (R$ {dados.at[indice, 'IPI']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_ipi != '':
        dados.at[indice, 'IPI'] = novo_valor_ipi

    if novo_fornecedor or nova_data_emissao or nova_data_pagamento or novo_valor != '' or novo_valor_icms != '' or novo_valor_cofins != '' or novo_valor_pis != '' or novo_valor_ipi != '':
        dados.to_csv(save_csv_file, index=False)
        print("Dados alterados com sucesso!")
    else:
        print("Nenhum dado foi alterado.")
    
    input('\nDigite uma tecla para voltar ao menu.')
    menu_alterar_dados()

# %%
def alterar_por_fornecedor():
    '''Altera os dados de uma despesa baseado no Fornecedor'''

    exibir_subtitulo("Alterar dados por Fornecedor")

    fornecedor_input = input("Insira o nome do Fornecedor que deseja alterar: ").strip()
    if not fornecedor_input:
        print("Valor inválido. Retornando ao menu...")
        menu_alterar_dados()
        return

    global dados
    dados_filtrados = dados[dados['Fornecedor'].str.contains(fornecedor_input, case=False, na=False)]

    if dados_filtrados.empty:
        print(f"Nenhum registro encontrado para o fornecedor '{fornecedor_input}'. Retornando ao menu...")
        menu_alterar_dados()
        return

    print(f"Registros encontrados para o fornecedor '{fornecedor_input}':")
    print(dados_filtrados)

    nNF_input = input("Insira o N° da NF que deseja alterar: ")
    try:
        nNF_input = int(nNF_input)
    except ValueError:
        print("Número da NF inválido. Retornando ao menu...")
        menu_alterar_dados()
        return

    if nNF_input not in dados_filtrados['N° da NF'].values:
        print(f"Número da NF {nNF_input} não encontrado para o fornecedor '{fornecedor_input}'. Retornando ao menu...")
        menu_alterar_dados()
        return

    indice = dados.index[dados['N° da NF'] == nNF_input][0]
    print(f"Alterando dados para a NF {nNF_input}:")
    
    novo_fornecedor = input(f"Fornecedor atual ({dados.at[indice, 'Fornecedor']}), novo valor (pressione Enter para manter): ")
    if novo_fornecedor:
        dados.at[indice, 'Fornecedor'] = novo_fornecedor

    nova_data_emissao = request_date("Emissão (pressione Enter para manter)")
    if nova_data_emissao:
        dados.at[indice, 'Data de emissão'] = nova_data_emissao

    nova_data_pagamento = request_date("Vencimento (pressione Enter para manter)")
    if nova_data_pagamento:
        dados.at[indice, 'Data de Vencimento'] = nova_data_pagamento

    novo_valor = input_valor_monetario(f"V. Total da NF atual (R$ {dados.at[indice, 'V. Total da NF']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor != '':
        dados.at[indice, 'V. Total da NF'] = novo_valor

    novo_valor_icms = input_valor_monetario(f"ICMS atual (R$ {dados.at[indice, 'ICMS']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_icms != '':
        dados.at[indice, 'ICMS'] = novo_valor_icms

    novo_valor_cofins = input_valor_monetario(f"COFINS atual (R$ {dados.at[indice, 'COFINS']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_cofins != '':
        dados.at[indice, 'COFINS'] = novo_valor_cofins

    novo_valor_pis = input_valor_monetario(f"PIS atual (R$ {dados.at[indice, 'PIS']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_pis != '':
        dados.at[indice, 'PIS'] = novo_valor_pis

    novo_valor_ipi = input_valor_monetario(f"IPI atual (R$ {dados.at[indice, 'IPI']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_ipi != '':
        dados.at[indice, 'IPI'] = novo_valor_ipi

    if novo_fornecedor or nova_data_emissao or nova_data_pagamento or novo_valor != '' or novo_valor_icms != '' or novo_valor_cofins != '' or novo_valor_pis != '' or novo_valor_ipi != '':
        dados.to_csv(save_csv_file, index=False)
        print("Dados alterados com sucesso!")
    else:
        print("Nenhum dado foi alterado.")
    
    input('\nDigite uma tecla para voltar ao menu.')
    menu_alterar_dados()

# %%
def alterar_por_data_vencimento():
    '''Altera os dados de uma despesa baseado na Data de Vencimento'''
    
    exibir_subtitulo("Alterar registro por Data de Vencimento")

    data_vencimento = request_date("Vencimento para busca")
    if data_vencimento is None:
        print("Data inválida. Retornando ao menu...")
        menu_alterar_dados()
        return

    global dados
    dados_filtrados = dados[dados['Data de Vencimento'] == str(data_vencimento)]

    if dados_filtrados.empty:
        print(f"Nenhum registro encontrado para a data de vencimento {data_vencimento}. Retornando ao menu...")
        menu_alterar_dados()
        return

    print(f"Registros encontrados para a data de vencimento {data_vencimento}:")
    print(dados_filtrados)

    nNF_input = input("Insira o N° da NF que deseja alterar: ")
    try:
        nNF_input = int(nNF_input)
    except ValueError:
        print("Número da NF inválido. Retornando ao menu...")
        menu_alterar_dados()
        return

    if nNF_input not in dados_filtrados['N° da NF'].values:
        print(f"Número da NF {nNF_input} não encontrado na data de vencimento {data_vencimento}. Retornando ao menu...")
        menu_alterar_dados()
        return

    indice = dados.index[dados['N° da NF'] == nNF_input][0]
    print(f"Alterando dados para a NF {nNF_input}:")
    
    novo_fornecedor = input(f"Fornecedor atual ({dados.at[indice, 'Fornecedor']}), novo valor (pressione Enter para manter): ")
    if novo_fornecedor:
        dados.at[indice, 'Fornecedor'] = novo_fornecedor

    nova_data_emissao = request_date("Emissão (pressione Enter para manter)")
    if nova_data_emissao:
        dados.at[indice, 'Data de emissão'] = nova_data_emissao

    nova_data_pagamento = request_date("Vencimento (pressione Enter para manter)")
    if nova_data_pagamento:
        dados.at[indice, 'Data de Vencimento'] = nova_data_pagamento

    novo_valor = input_valor_monetario(f"V. Total da NF atual (R$ {dados.at[indice, 'V. Total da NF']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor != '':
        dados.at[indice, 'V. Total da NF'] = novo_valor

    novo_valor_icms = input_valor_monetario(f"ICMS atual (R$ {dados.at[indice, 'ICMS']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_icms != '':
        dados.at[indice, 'ICMS'] = novo_valor_icms

    novo_valor_cofins = input_valor_monetario(f"COFINS atual (R$ {dados.at[indice, 'COFINS']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_cofins != '':
        dados.at[indice, 'COFINS'] = novo_valor_cofins

    novo_valor_pis = input_valor_monetario(f"PIS atual (R$ {dados.at[indice, 'PIS']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_pis != '':
        dados.at[indice, 'PIS'] = novo_valor_pis

    novo_valor_ipi = input_valor_monetario(f"IPI atual (R$ {dados.at[indice, 'IPI']}), novo valor (pressione Enter para manter): R$ ")
    if novo_valor_ipi != '':
        dados.at[indice, 'IPI'] = novo_valor_ipi

    if novo_fornecedor or nova_data_emissao or nova_data_pagamento or novo_valor != '' or novo_valor_icms != '' or novo_valor_cofins != '' or novo_valor_pis != '' or novo_valor_ipi != '':
        dados.to_csv(save_csv_file, index=False)
        print("Dados alterados com sucesso!")
    else:
        print("Nenhum dado foi alterado.")
    
    input('\nDigite uma tecla para voltar ao menu.')
    menu_alterar_dados()

# %%
def apagar_por_nNF():
    '''Apaga os dados de uma despesa baseado no N° da NF'''

    exibir_subtitulo("Apagar registro por N° da NF")

    nNF_input = input("Insira o N° da NF que deseja apagar: ")
    try:
        nNF_input = int(nNF_input)
    except ValueError:
        print("Número da NF inválido. Retornando ao menu...")
        menu_alterar_dados()
        return

    global dados
    if nNF_input not in dados['N° da NF'].values:
        print(f"Número da NF {nNF_input} não encontrado. Retornando ao menu...")
        menu_alterar_dados()
        return
    elif len(dados[dados['N° da NF'] == nNF_input]) > 1:
        print(f"Mais de um registro encontrado para a NF {nNF_input}.\n")
        print(dados[dados['N° da NF'] == nNF_input])
        index = input("Insira o índice do registro que deseja apagar (conforme exibido acima): ")
        try:
            index = int(index)
            if index not in dados.index:
                print("Índice inválido. Retornando ao menu...")
                menu_alterar_dados()
                return
        except ValueError:
            print("Valor inserido inválido. Retornando ao menu...")
            menu_alterar_dados()
            return
    else:
        index = dados.index[dados['N° da NF'] == nNF_input][0]

    confirmacao = input(f"Tem certeza que deseja apagar a NF {nNF_input}? (s/n): ").strip().lower()
    if confirmacao == 's':
        dados = dados.drop(index)
        dados.to_csv(save_csv_file, index=False)
        print("Registro apagado com sucesso!")
    else:
        print("Operação cancelada.")

    input('\nDigite uma tecla para voltar ao menu.')
    menu_alterar_dados()

# %%
def exibir_opcoes_alterar():
    '''Exibe todas as opcoes de input do usuario no menu alterar dados'''
    print('1. Alterar Gastos - Procurar por N° da NF')
    print('2. Alterar Gastos - Procurar por Fornecedor')
    print('3. Alterar Gastos - Procurar por Data de Emissão')
    print('4. Alterar Gastos - Procurar por Data de Vencimento')
    print('5. Apagar Gastos - Procurar por N° da NF')
    print('9. Voltar ao Menu Principal\n')

def escolher_opcao_alterar():
    ''' Solicita a executa a opcao escolhida pelo usuario
    
    Outputs:
    -Executa a opcao escolhida pelo usuario
    '''
    try:
        opcao_escolhida = int(input('Escolha uma opção: '))
        # opcao_escolhida = int(opcao_escolhida)

        if opcao_escolhida == 1: 
            alterar_por_nNF()
        elif opcao_escolhida == 2: 
            alterar_por_fornecedor()
        elif opcao_escolhida == 3: 
            alterar_por_data_emissao()            
        elif opcao_escolhida == 4:
            alterar_por_data_vencimento()
        elif opcao_escolhida == 5:
            apagar_por_nNF()
        elif opcao_escolhida == 9: 
            voltar_ao_menu_principal()
        else: 
            opcao_invalida()
    except Exception as e:
        if debug:
            print(f'Error: {type(e).__name__}')
            print(f"Error message: {e}")
        opcao_invalida()

def menu_alterar_dados():
    '''Menu para alterar dados de jogadoras'''
    os.system('cls')
    exibir_titulo(alterar_titulo)
    exibir_opcoes_alterar()
    escolher_opcao_alterar()

# %% [markdown]
# #### debug

# %%
def alternar_debug():
    '''Alterna o modo debug'''
    global debug
    global save_csv_file
    debug = not debug
    status = 'ativado' if debug else 'desativado'
    exibir_subtitulo('Alterar Modo Debug')
    print(f'\nModo debug {status}.')
    if debug:
        save_csv_file = 'src/Data/Data_Debug.csv'
    else:
        save_csv_file = './src/Data/Dados.csv'
    time.sleep(1)
    voltar_ao_menu_principal()

# %% [markdown]
# ## Principal

# %%
def exibir_opcoes():
    '''Exibe todas as opcoes de input do usuario'''
    print('1. Registrar Custos')
    print('2. Visualizar Despesas')
    print('3. Alterar Dados')
    print('4. Alternar Debug')
    print('9. Sair\n')

def escolher_opcao():
    ''' Solicita a executa a opcao escolhida pelo usuario
    
    Outputs:
    -Executa a opcao escolhida pelo usuario
    '''
    try:
        opcao_escolhida = int(input('Escolha uma opção: '))
        # opcao_escolhida = int(opcao_escolhida)

        if opcao_escolhida == 1: 
            menu_registrar_dados()
        elif opcao_escolhida == 2: 
            menu_visualizar_dados()
        elif opcao_escolhida == 3: 
            menu_alterar_dados()            
        elif opcao_escolhida == 4:
            alternar_debug()
        elif opcao_escolhida == 9: 
            finalizar_app()
        else: 
            opcao_invalida()
    except Exception as e:
        if debug:
            print(f'Error: {type(e).__name__}')
            print(f"Error message: {e}")
        opcao_invalida()

def main_menu():
    '''Funcao principal que inicial o programa'''
    os.system('cls')    
    exibir_titulo(titulo_empresa)
    exibir_opcoes()
    escolher_opcao()


# %% [markdown]
# # Main

# %%
def main():
    carregar_dados()
    main_menu()

if __name__ == '__main__':
    main()


