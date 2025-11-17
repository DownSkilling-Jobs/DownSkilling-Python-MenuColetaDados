from openpyxl.utils.dataframe import dataframe_to_rows
import datetime as dt
import openpyxl as xl
import pandas as pd

TabList = ['JAN', 'FEV', 'MAR', 'ABR', 'MAI', 'JUN', 'JUL', 'AGO', 'SET', 'OUT', 'NOV', 'DEZ']
today = dt.date.today()
current_month_index = today.month - 1

def merge_csv_to_excel(csv_file, excel_file, output_file):
    # Load the existing Excel workbook
    workbook = xl.load_workbook(excel_file)
    sheet = workbook[TabList[current_month_index]]

    style = xl.styles.NamedStyle(name="standard_style")
    style.font = xl.styles.Font(name='Aptos', size=11)
    style.alignment = xl.styles.Alignment(horizontal='center', vertical='center')
    style.border = xl.styles.Border(
        left=xl.styles.Side(border_style='medium', color='000000'),
        right=xl.styles.Side(border_style='medium', color='000000'),
        top=xl.styles.Side(border_style='none', color='000000'),
        bottom=xl.styles.Side(border_style='none', color='000000')
    )
    workbook.add_named_style(style)

    # Read the CSV file into a DataFrame
    df = pd.read_csv(csv_file)

    # Append the DataFrame to the Excel sheet
    for r in dataframe_to_rows(df, index=False, header=False):
        if 'current_row' not in locals():
            current_row = 3
        for col_idx, value in enumerate(r, start=1):
            sheet.cell(row=current_row, column=col_idx, value=value)
            sheet.cell(row=current_row, column=col_idx).style = "standard_style"
            if col_idx in [ 5, 7, 8, 9, 10 ]:  # Columns with monetary values
                sheet.cell(row=current_row, column=col_idx).number_format = 'R$ #,##0.00'
            
        current_row += 1

    # Save the updated workbook to a new file
    workbook.save(output_file)

if __name__ == "__main__":
    csv_file = 'src/Data/Data2.csv'
    excel_file = 'src/template/DM-2025 - Template.xlsx'
    output_file = 'src/Data/Merged_Output_test.xlsx'
    merge_csv_to_excel(csv_file, excel_file, output_file)

    